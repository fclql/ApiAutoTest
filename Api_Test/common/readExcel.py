# -*- coding: utf-8 -*-
import openpyxl


class Case:
    def __init__(self):
        self.case_id = None
        self.url = None
        self.data = None
        self.title = None
        self.method = None
        self.excepted = None


class ReadExcel:
    def __init__(self,filepath):
        try:
            self.workbook = openpyxl.load_workbook(filepath)
        except FileNotFoundError as e:
            print("{0}没有找到，请检查路径".format(filepath))
            raise e

    def get_cases(self,sheetname):
        sheetname = self.workbook[sheetname]
        sheet_max_row = sheetname.max_row
        cases = []
        for i in range(2,sheet_max_row+1):
            case = Case()
            case.case_id = sheetname.cell(i, 1).value
            case.url = sheetname.cell(i, 2).value
            case.data = sheetname.cell(i, 3).value
            case.title = sheetname.cell(i, 4).value
            case.method = sheetname.cell(i, 5).value
            case.excepted = sheetname.cell(i, 6).value
            cases.append(case)
        return cases
